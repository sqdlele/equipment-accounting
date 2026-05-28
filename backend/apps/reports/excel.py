import openpyxl

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from openpyxl.utils import get_column_letter

from datetime import date





def _thin_border():

    thin = Side(style='thin')

    return Border(left=thin, right=thin, top=thin, bottom=thin)





def _header_fill():

    return PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')





def generate_asset_excel(assets):

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = 'Инвентаризационная ведомость'



    num_cols = 14

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)

    title_cell = ws['A1']

    title_cell.value = 'ИНВЕНТАРИЗАЦИОННАЯ ВЕДОМОСТЬ ТЕХНИКИ'

    title_cell.font = Font(bold=True, size=14, color='FFFFFF')

    title_cell.fill = _header_fill()

    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 30



    ws['A2'] = f'Дата формирования: {date.today().strftime("%d.%m.%Y")}'

    ws['A2'].font = Font(italic=True)

    ws.row_dimensions[2].height = 20



    headers = [

        '№', 'Инв. номер', 'Наименование', 'Модель', 'Категория',

        'Производитель', 'Статус', 'Локация', 'Ответственный',

        'Дата покупки', 'Год изг.', 'Серийный номер', 'Стоимость', 'Описание',

    ]

    header_row = 3

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=header_row, column=col_num, value=header)

        cell.font = Font(bold=True, color='FFFFFF')

        cell.fill = _header_fill()

        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        cell.border = _thin_border()

    ws.row_dimensions[header_row].height = 25



    status_colors = {

        'На складе': 'D1FAE5',

        'В эксплуатации': 'DBEAFE',

        'В ремонте': 'FEF3C7',

        'Списан': 'FEE2E2',

    }



    for row_num, asset in enumerate(assets, 1):

        row = header_row + row_num

        status_display = asset.get_status_display()

        desc = (asset.description or '')[:500]

        row_data = [

            row_num,

            asset.inventory_number,

            asset.name,

            asset.model or '',

            asset.category.name if asset.category else '',

            asset.manufacturer.name if asset.manufacturer else '',

            status_display,

            asset.location.name if asset.location else '',

            asset.responsible_employee.full_name if asset.responsible_employee else '',

            asset.purchase_date.strftime('%d.%m.%Y') if asset.purchase_date else '',

            asset.manufacture_year or '',

            asset.serial_number or '',

            float(asset.price) if asset.price is not None else '',

            desc,

        ]

        fill_color = status_colors.get(status_display, 'FFFFFF')

        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')



        for col_num, value in enumerate(row_data, 1):

            cell = ws.cell(row=row, column=col_num, value=value)

            cell.border = _thin_border()

            cell.fill = fill

            cell.alignment = Alignment(vertical='center', wrap_text=True)

        ws.row_dimensions[row].height = 18



    col_widths = [5, 15, 28, 22, 18, 16, 14, 18, 22, 12, 10, 18, 12, 36]

    for i, width in enumerate(col_widths, 1):

        ws.column_dimensions[get_column_letter(i)].width = width



    ws.freeze_panes = 'A4'

    return wb





def generate_repair_excel(tickets):

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = 'Журнал ремонтов'



    ws.merge_cells('A1:I1')

    title_cell = ws['A1']

    title_cell.value = 'ЖУРНАЛ РЕМОНТОВ И ТЕХНИЧЕСКОГО ОБСЛУЖИВАНИЯ'

    title_cell.font = Font(bold=True, size=14, color='FFFFFF')

    title_cell.fill = _header_fill()

    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 30



    ws['A2'] = f'Дата формирования: {date.today().strftime("%d.%m.%Y")}'

    ws['A2'].font = Font(italic=True)



    headers = ['№', 'Актив', 'Инв. номер', 'Статус', 'Приоритет',

               'Описание дефекта', 'Исполнитель', 'Создан', 'Решён', 'Простой (ч.)']

    for col_num, header in enumerate(headers, 1):

        cell = ws.cell(row=3, column=col_num, value=header)

        cell.font = Font(bold=True, color='FFFFFF')

        cell.fill = _header_fill()

        cell.alignment = Alignment(horizontal='center', vertical='center')

        cell.border = _thin_border()

    ws.row_dimensions[3].height = 25



    for row_num, ticket in enumerate(tickets, 1):

        row = 3 + row_num

        row_data = [

            row_num,

            ticket.asset.name,

            ticket.asset.inventory_number,

            ticket.get_status_display(),

            ticket.get_priority_display(),

            ticket.description[:100],

            ticket.assigned_to.full_name if ticket.assigned_to else '',

            ticket.created_at.strftime('%d.%m.%Y %H:%M'),

            ticket.resolved_at.strftime('%d.%m.%Y %H:%M') if ticket.resolved_at else '',

            ticket.downtime_hours,

        ]

        for col_num, value in enumerate(row_data, 1):

            cell = ws.cell(row=row, column=col_num, value=value)

            cell.border = _thin_border()

            cell.alignment = Alignment(vertical='center')

        ws.row_dimensions[row].height = 18



    col_widths = [5, 28, 15, 15, 12, 40, 25, 18, 18, 12]

    for i, width in enumerate(col_widths, 1):

        ws.column_dimensions[get_column_letter(i)].width = width



    ws.freeze_panes = 'A4'

    return wb

