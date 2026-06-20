"""
Импорт активов из Excel (.xlsx). Строка заголовков как в экспортируемой ведомости или шаблоне.
"""
import openpyxl
from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.db import transaction

from apps.employees.models import Employee
from apps.locations.models import Location
from .models import Asset, Category, Manufacturer

HEADER_ALIASES = {
    'инв. номер': 'inventory_number',
    'инвентарный номер': 'inventory_number',
    'наименование': 'name',
    'модель': 'model',
    'категория': 'category',
    'производитель': 'manufacturer',
    'статус': 'status',
    'локация': 'location',
    'ответственный': 'responsible',
    'дата покупки': 'purchase_date',
    'год изг.': 'manufacture_year',
    'год изготовления': 'manufacture_year',
    'серийный номер': 'serial_number',
    'стоимость': 'price',
    'описание': 'description',
}

IMPORT_TEMPLATE_HEADERS = [
    'Инв. номер',
    'Наименование',
    'Модель',
    'Категория',
    'Производитель',
    'Статус',
    'Локация',
    'Ответственный',
    'Дата покупки',
    'Год изг.',
    'Серийный номер',
    'Стоимость',
    'Описание',
]


def build_import_template_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Импорт'
    ws['A1'] = 'Шаблон импорта ИТ-активов'
    ws['A2'] = (
        'Статус: На складе | В эксплуатации | В ремонте | Списан. '
        'Локация и ответственный - как в справочниках. Даты: ДД.ММ.ГГГГ'
    )
    for i, h in enumerate(IMPORT_TEMPLATE_HEADERS, start=1):
        ws.cell(row=3, column=i, value=h)
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 22
    return wb


STATUS_LABEL_TO_CODE = {
    'на складе': Asset.STATUS_WAREHOUSE,
    'в эксплуатации': Asset.STATUS_IN_USE,
    'в ремонте': Asset.STATUS_REPAIR,
    'списан': Asset.STATUS_WRITTEN_OFF,
}


def _normalize_header(cell_value) -> str | None:
    if cell_value is None:
        return None
    s = str(cell_value).strip().lower()
    return s or None


def _find_header_row(ws, max_scan=15):
    for r in range(1, max_scan + 1):
        row_vals = [_normalize_header(ws.cell(row=r, column=c).value) for c in range(1, ws.max_column + 1)]
        if any(v in ('инв. номер', 'инвентарный номер') for v in row_vals if v):
            mapping = {}
            for c, v in enumerate(row_vals, start=1):
                if not v:
                    continue
                key = HEADER_ALIASES.get(v)
                if key:
                    mapping[c] = key
            if 'inventory_number' in mapping.values():
                inv_col = next(k for k, v in mapping.items() if v == 'inventory_number')
                return r, mapping, inv_col
    return None, {}, None


def _parse_date(val):
    if val is None or val == '':
        return None
    if hasattr(val, 'date') and not isinstance(val, datetime):
        return val  # date object from openpyxl
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_int(val):
    if val is None or val == '':
        return None
    try:
        return int(float(str(val).strip().replace(',', '.')))
    except (ValueError, TypeError):
        return None


def _parse_decimal(val):
    if val is None or val == '':
        return None
    try:
        return Decimal(str(val).strip().replace(' ', '').replace(',', '.'))
    except (InvalidOperation, ValueError):
        return None


def _resolve_category(name: str | None):
    if not name or not str(name).strip():
        return None
    cat, _ = Category.objects.get_or_create(name=str(name).strip()[:100], defaults={'description': ''})
    return cat


def _resolve_manufacturer(name: str | None):
    if not name or not str(name).strip():
        return None
    man, _ = Manufacturer.objects.get_or_create(name=str(name).strip()[:100])
    return man


def _resolve_location(name: str | None):
    if not name or not str(name).strip():
        return None
    q = str(name).strip()
    loc = Location.objects.filter(name__iexact=q).first()
    return loc


def _resolve_employee(name: str | None):
    if not name or not str(name).strip():
        return None
    q = str(name).strip()
    emp = Employee.objects.filter(full_name__iexact=q).first()
    if not emp and '@' in q:
        emp = Employee.objects.filter(email__iexact=q).first()
    return emp


def _resolve_status(raw) -> str | None:
    if raw is None or str(raw).strip() == '':
        return None
    s = str(raw).strip().lower()
    if s in STATUS_LABEL_TO_CODE:
        return STATUS_LABEL_TO_CODE[s]
    codes = {code for code, _ in Asset.STATUS_CHOICES}
    if s in codes:
        return s
    return None


def import_assets_workbook(wb) -> dict:
    ws = wb.active
    header_row, col_map, _ = _find_header_row(ws)
    if not header_row:
        return {
            'created': 0,
            'updated': 0,
            'errors': [{'row': None, 'message': 'Не найдена строка заголовков с колонкой «Инв. номер»'}],
        }

    errors = []
    created = 0
    updated = 0

    with transaction.atomic():
        for r in range(header_row + 1, ws.max_row + 1):
            row_data = {}
            for col, key in col_map.items():
                cell = ws.cell(row=r, column=col)
                val = cell.value
                if val is not None and isinstance(val, str):
                    val = val.strip()
                    if val == '':
                        val = None
                row_data[key] = val

            inv = row_data.get('inventory_number')
            if inv is None or str(inv).strip() == '':
                continue
            inventory_number = str(inv).strip()[:50]

            name = row_data.get('name')
            if not name or str(name).strip() == '':
                errors.append({'row': r, 'message': f'{inventory_number}: не указано наименование'})
                continue

            kwargs = {
                'serial_number': str(row_data.get('serial_number') or '')[:100],
                'name': str(name).strip()[:255],
                'model': str(row_data.get('model') or '').strip()[:255],
            }
            pd = _parse_date(row_data.get('purchase_date'))
            if pd:
                kwargs['purchase_date'] = pd
            my = _parse_int(row_data.get('manufacture_year'))
            if my:
                kwargs['manufacture_year'] = my if 1990 <= my <= 2100 else None
            pr = _parse_decimal(row_data.get('price'))
            if pr is not None:
                kwargs['price'] = pr
            desc = row_data.get('description')
            if desc:
                kwargs['description'] = str(desc)

            cat = _resolve_category(row_data.get('category'))
            if cat:
                kwargs['category'] = cat
            man = _resolve_manufacturer(row_data.get('manufacturer'))
            if man:
                kwargs['manufacturer'] = man

            st = _resolve_status(row_data.get('status'))
            if st:
                kwargs['status'] = st

            loc = _resolve_location(row_data.get('location'))
            if loc:
                kwargs['location'] = loc
            elif row_data.get('location'):
                errors.append({'row': r, 'message': f'{inventory_number}: локация «{row_data.get("location")}» не найдена'})
                continue

            emp = _resolve_employee(row_data.get('responsible'))
            if emp:
                kwargs['responsible_employee'] = emp
            elif row_data.get('responsible'):
                errors.append({'row': r, 'message': f'{inventory_number}: сотрудник «{row_data.get("responsible")}» не найден'})
                continue

            asset = Asset.objects.filter(inventory_number=inventory_number).first()
            if asset:
                for field, value in kwargs.items():
                    setattr(asset, field, value)
                asset.save()
                updated += 1
            else:
                Asset.objects.create(inventory_number=inventory_number, **kwargs)
                created += 1

    return {'created': created, 'updated': updated, 'errors': errors}
